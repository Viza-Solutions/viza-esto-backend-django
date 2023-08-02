from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from .models import *
from .serializers import *
from django.db import IntegrityError
from datetime import datetime
from uuid import uuid4
from io import BytesIO
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from django.http import HttpResponse, JsonResponse
import openpyxl

from sms_api.functions import *


# Views
@api_view(["GET"])
def payment_method_list(request):
    payment_methods = PaymentMethod.objects.filter(deleted=False).order_by("name")
    serializer = PaymentMethodSerializer(payment_methods, many=True)
    return Response(serializer.data)


@api_view(["GET"])
def client_payment_method_list(request, client_id):
    payment_methods = PaymentMethod.objects.filter(
        deleted=False, client_id=client_id
    ).order_by("name")
    serializer = PaymentMethodSerializer(payment_methods, many=True)
    return Response(serializer.data)


@api_view(["POST"])
def create_payment_method(request):
    serializer = PaymentMethodSerializer(data=request.data)
    if serializer.is_valid():
        try:
            serializer.save(added_by=request.user)
            return Response(
                {
                    "message": "Payment method created successfully",
                    "data": serializer.data,
                },
                status=status.HTTP_201_CREATED,
            )
        except IntegrityError:
            return Response(
                {
                    "error": "A payment method with the same name already exists for this client."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["GET"])
def retrieve_payment_method(request, pk):
    try:
        payment_method = PaymentMethod.objects.get(pk=pk, deleted=False)
        serializer = PaymentMethodSerializer(payment_method)
        return Response(serializer.data)
    except PaymentMethod.DoesNotExist:
        return Response(
            {"message": "Payment method not found"}, status=status.HTTP_404_NOT_FOUND
        )


@api_view(["PUT"])
def update_payment_method(request, pk):
    try:
        payment_method = PaymentMethod.objects.get(pk=pk, deleted=False)
        serializer = PaymentMethodSerializer(
            payment_method, data=request.data, partial=True
        )
        if serializer.is_valid():
            try:
                serializer.save()
                return Response(
                    {
                        "message": "Payment method updated successfully",
                        "data": serializer.data,
                    }
                )
            except IntegrityError:
                return Response(
                    {
                        "error": "A payment method with the same name already exists for this client."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    except PaymentMethod.DoesNotExist:
        return Response(
            {"message": "Payment method not found"}, status=status.HTTP_404_NOT_FOUND
        )


@api_view(["DELETE"])
def delete_payment_method(request, pk):
    try:
        payment_method = PaymentMethod.objects.get(pk=pk, deleted=False)
        payment_method.deleted = True
        payment_method.save()
        return Response(
            {"message": "Payment method deleted successfully"},
            status=status.HTTP_204_NO_CONTENT,
        )
    except PaymentMethod.DoesNotExist:
        return Response(
            {"message": "Payment method not found"}, status=status.HTTP_404_NOT_FOUND
        )


# PAYMENT
@api_view(["POST"])
def create_payment_transaction(request):
    serializer = PaymentTransactionSerializer(data=request.data)
    if serializer.is_valid():
        tenant_id = serializer.validated_data["tenant"].id

        # Retrieve tenant and room price in a single query
        tenant = Tenant.objects.select_related("room", "client").get(pk=tenant_id)
        room_price = tenant.room.monthly_price

        current_month = datetime.utcnow().month
        current_year = datetime.utcnow().year

        payment_transactions = PaymentTransaction.objects.filter(
            tenant=tenant
        ).order_by("-created_at")

        if payment_transactions.exists():
            first_payment_transaction = payment_transactions.first()
            last_paid_month = first_payment_transaction.month
            last_paid_year = first_payment_transaction.year

            unpaid_months = (
                (current_year - last_paid_year) * 12 + current_month - last_paid_month
            )
            rent_to_be_paid = unpaid_months * room_price

            new_balance = serializer.validated_data["amount"] - (
                rent_to_be_paid - first_payment_transaction.balance
            )
        else:
            new_balance = serializer.validated_data["amount"] - room_price

        # Validate the client ID
        client_id = serializer.validated_data["client"].id
        if tenant.client_id != client_id:
            return Response(
                {
                    "error": "Invalid client id provided. The client does not match the tenant."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Generate a unique UUID
        uuid = str(uuid4())

        serializer.save(
            balance=new_balance,
            month=current_month,
            year=current_year,
            uuid=uuid,
        )

        # send sms
        amount_paid = serializer.validated_data["amount"]
        current_datetime_sms = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        # message = f"Thank you for your payment of KES {amount_paid:,.2f}. Your rent has been received. Your current balance is KES {new_balance:,.2f} as of {current_datetime_sms}."

        if new_balance > 0:
            # overpayment
            mm = f" You have made an overpayment of KES {new_balance:,.2f}"
        elif new_balance < 0:
            # underpayment
            mm = f" Your current balance after the payment is KES {abs(new_balance):,.2f}"
        else:
            # settled
            mm = " Your account has been settled fully"
        message = (
            f"Dear {tenant.fullname},\n\n"
            f"Your rent payment of KES {amount_paid:,.2f} has been successfully received. "
            f"{mm} as of {current_datetime_sms}.\n\n"
            # f"Should you have any queries or require further assistance, please don't hesitate to reach out. Thank you for choosing us as your property management team.\n\n"
            f"Best regards,\n"
            f"Your Property Management Team"
        )
        recepient = [tenant.phone_number]
        send_sms(message, recepient)

        return Response(
            {
                "message": "Payment transaction created successfully",
                "receipt_number": serializer.data["id"],
                "balance": serializer.data["balance"],
                # all the data
                # "data": serializer.data,
            },
            status=status.HTTP_201_CREATED,
        )
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


@api_view(["DELETE"])
def delete_all_payment_transactions(request):
    # Delete all payment transactions
    PaymentTransaction.objects.all().delete()
    return Response(
        {"message": "All payment transactions deleted successfully"},
        status=status.HTTP_204_NO_CONTENT,
    )


# reports


@api_view(["GET"])
def excel_report_view(request, tenant_id):
    try:
        # Retrieve the data from the PaymentTransaction model for the specific Tenant
        queryset = PaymentTransaction.objects.filter(
            tenant_id=tenant_id, reversed=False
        )

        if not queryset.exists():
            return JsonResponse(
                {"error": "No transactions found for the given tenant_id."}, status=404
            )

        # Create the Excel file
        workbook = openpyxl.Workbook()
        worksheet = workbook.active

        # Write the header row with bold font, centered alignment, and font color
        header_font = openpyxl.styles.Font(bold=True)
        header_alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )

        headers = [
            "ID",
            "Amount",
            "Balance",
            "Month",
            "Year",
            "Payment Method",
            "Reference",
            "Description",
            "Processed By",
            "Client",
            "Created At",
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.value = header.upper()  # Capitalize the header
            cell.font = header_font
            cell.alignment = header_alignment

        # Write data rows
        data_alignment = openpyxl.styles.Alignment(
            horizontal="center", vertical="center"
        )
        for row_num, transaction in enumerate(queryset, start=2):
            data = [
                transaction.id,
                transaction.amount,
                transaction.balance,
                transaction.month,
                transaction.year,
                str(transaction.payment_method),
                transaction.reference,
                transaction.description,
                str(transaction.processed_by),
                str(transaction.client),
                transaction.created_at,
            ]
            for col_num, value in enumerate(data, start=1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.alignment = data_alignment

        # Get the current date and time
        current_datetime = datetime.now()

        # Get the current month and year
        current_month = int(current_datetime.month)
        current_year = int(current_datetime.year)

        # Calculate unpaid and prepaid months
        last_transaction = queryset.order_by("-id").first()
        balance = int(last_transaction.balance)
        year = int(last_transaction.year)
        month = int(last_transaction.month)

        try:
            # Find the room with the given tenant_id
            tenant_details = Tenant.objects.get(pk=tenant_id)
            room_id = tenant_details.room_id
            name = tenant_details.fullname

            # Get the room object using the retrieved room_id
            room = Room.objects.get(pk=room_id)
            monthly_price = int(room.monthly_price)

        except Room.DoesNotExist:
            # Handle the case when no room is found for the given tenant_id
            room_id = None
            monthly_price = None
            name = "Undefined"

        worksheet.title = f"Statement- {name}"

        # Calculate the curr_balance
        months_difference = ((current_year - year) * 12) + (current_month - month)

        curr_balance = (-months_difference * monthly_price) + balance

        over_payment = 0
        under_payment = 0

        if curr_balance > 0:
            over_payment = curr_balance
        else:
            under_payment = curr_balance

        # Write unpaid and prepaid months information at the end of the sheet
        unpaid_cell = worksheet.cell(
            row=len(queryset) + 5, column=1, value="Unpaid Amount"
        )
        unpaid_cell.font = openpyxl.styles.Font(bold=True)
        unpaid_cell.alignment = openpyxl.styles.Alignment(horizontal="right")

        unpaid_value_cell = worksheet.cell(
            row=len(queryset) + 5, column=2, value=under_payment
        )
        unpaid_value_cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        prepaid_cell = worksheet.cell(
            row=len(queryset) + 6, column=1, value="Prepaid Amount"
        )
        prepaid_cell.font = openpyxl.styles.Font(bold=True)
        prepaid_cell.alignment = openpyxl.styles.Alignment(horizontal="right")

        prepaid_value_cell = worksheet.cell(
            row=len(queryset) + 6, column=2, value=over_payment
        )
        prepaid_value_cell.alignment = openpyxl.styles.Alignment(horizontal="center")

        # Auto-fit column width for all columns
        for col_num, header in enumerate(headers, start=1):
            column_letter = openpyxl.utils.get_column_letter(col_num)
            worksheet.column_dimensions[column_letter].auto_size = True

        # Create a response with the Excel file
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        current_datetime = datetime.now()

        # Generate the timestamp in the Nairobi timezone
        timestamp = current_datetime.strftime("%Y%m%d_%H%M%S")

        # Append the timestamp to the filename
        filename = f"{name}_Rent_Report_{timestamp}.xlsx"

        # Set the Content-Disposition header with the updated filename
        response["Content-Disposition"] = f"attachment; filename={filename}"
        workbook.save(response)

        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["GET"])
def pdf_report_view(request, tenant_id):
    try:
        # Retrieve the data from the PaymentTransaction model for the specific Tenant
        queryset = PaymentTransaction.objects.filter(
            tenant_id=tenant_id, reversed=False
        )

        if not queryset.exists():
            return JsonResponse(
                {"error": "No transactions found for the given tenant_id."}, status=404
            )

        # Define styles for the header and other text
        styles = getSampleStyleSheet()
        header_style = ParagraphStyle(
            "Header",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            alignment=1,  # 0 for left, 1 for center, 2 for right
            fontSize=18,
        )
        info_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=14,
            textColor="black",
            alignment=1,
        )
        tenant_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=12,
            textColor="black",
            alignment=0,
        )

        bal_style = ParagraphStyle(
            "Info",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            textColor="black",
            alignment=2,
        )

        # Create the PDF document
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))

        # Create the table and add data
        data = [
            [
                "Transaction No.",
                "Amount",
                "Balance",
                "Month",
                "Year",
                "Payment Method",
                "Reference",
                "Description",
                "Processed By",
                "Client",
                "Processed On",
            ]
        ]
        for transaction in queryset:
            data.append(
                [
                    transaction.id,
                    transaction.amount,
                    transaction.balance,
                    transaction.month,
                    transaction.year,
                    str(transaction.payment_method),
                    transaction.reference,
                    transaction.description,
                    str(transaction.processed_by),
                    str(transaction.client),
                    transaction.created_at,
                ]
            )

        # Add table to the PDF
        table = Table(data)
        table.setStyle(
            TableStyle(
                [
                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.grey,
                    ),  # Header row background color
                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.whitesmoke,
                    ),  # Header row text color
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),  # Center align all cells
                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold",
                    ),  # Bold font for header row
                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, 0),
                        12,
                    ),  # Add padding to the header row
                    (
                        "BACKGROUND",
                        (0, 1),
                        (-1, -1),
                        "#ffffff",
                    ),  # Background color for data rows
                ]
            )
        )

        # Add the heading to the top of the page with more information
        title_paragraph = Paragraph("KEIGHT", header_style)

        additional_info = Paragraph(
            "PO Box 21,<br/>"
            "Nairobi, Kenya<br/>"
            "Tel: +254 58 033088<br/>"
            "Email: info@keights.com<br/>",
            info_style,
        )

        # doc.topMargin = 70  # Adjust the top margin to make space for the header
        # doc.build([title_paragraph, additional_info, Spacer(1, 20), table])

        # Create a response with the PDF file
        response = HttpResponse(content_type="application/pdf")

        current_datetime = datetime.now()

        # Generate the timestamp in the Nairobi timezone
        timestamp = current_datetime.strftime("%Y%m%d_%H%M%S")

        # Find the room with the given tenant_id (You may need to implement the Tenant and Room models)

        # Get the current date and time
        current_datetime = datetime.now()

        # Get the current month and year
        current_month = int(current_datetime.month)
        current_year = int(current_datetime.year)

        # Calculate unpaid and prepaid months
        last_transaction = queryset.order_by("-id").first()
        balance = int(last_transaction.balance)
        year = int(last_transaction.year)
        month = int(last_transaction.month)

        try:
            tenant_details = Tenant.objects.get(pk=tenant_id)
            room_id = tenant_details.room_id
            name = (
                tenant_details.fullname
            )  # Set the 'name' variable with the tenant's name

            # Get the room object using the retrieved room_id
            room = Room.objects.get(pk=room_id)
            room_number = room.room_number
            estate = str(room.property)
            monthly_price = int(room.monthly_price)

        except Tenant.DoesNotExist:
            # Handle the case when no tenant is found for the given tenant_id
            room_id = None
            monthly_price = None
            name = "Undefined"

        # Calculate the curr_balance
        months_difference = ((current_year - year) * 12) + (current_month - month)

        curr_balance = (-months_difference * monthly_price) + balance

        if curr_balance > 0:
            curr_balance_str = "Prepaid Amount Ksh. " + str(curr_balance) + "/="
        else:
            curr_balance_str = "Underpaid Amount Ksh. " + str(curr_balance) + "/="

        tenant_info = Paragraph(
            "<br/>"
            "Name : " + name + "<br/>"
            "Estate : " + estate + "<br/>"
            "Room No : " + str(room_number) + "<br/>"
            "Rent : Ksh. " + str(monthly_price) + "/=",
            tenant_style,
        )

        balance_info = Paragraph(
            "<br/>" "<br/>" " " + str(curr_balance_str),
            bal_style,
        )
        doc.topMargin = 70  # Adjust the top margin to make space for the header
        doc.build(
            [
                title_paragraph,
                additional_info,
                tenant_info,
                Spacer(1, 20),
                table,
                balance_info,
            ]
        )

        # Append the timestamp to the filename
        filename = f"{name}_Rent_Report_{timestamp}.pdf"

        # Set the Content-Disposition header with the updated filename
        response["Content-Disposition"] = f"attachment; filename={filename}"
        response.write(buffer.getvalue())

        return response

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)


# get all tenants transactions


@api_view(["GET"])
def get_transactions_for_tenant(request, tenant_id):
    try:
        # Retrieve all transactions for the specific tenant where reversed=False
        transactions = PaymentTransaction.objects.filter(
            tenant_id=tenant_id, reversed=False
        ).order_by("-id")

        # Serialize the transactions
        serializer = PaymentTransactionSerializer(transactions, many=True)

        return Response(serializer.data)

    except PaymentTransaction.DoesNotExist:
        return Response(
            {"error": "No transactions found for the given tenant_id."}, status=404
        )

    except Exception as e:
        return Response({"error": str(e)}, status=500)
